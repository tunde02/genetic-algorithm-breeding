import json, pickle
from openpyxl import load_workbook
from data_loader import Point, RedPoint
from haversine import haversine


point_list = []
red_point_list = []
red_point_index_list = []

point_data_wb = load_workbook('./data/poi_coord.xlsx')
point_data_sh = point_data_wb.active

# 1. point_list 불러오기
for row in range(2, 3000 + 2):
    point_index = point_data_sh[f'Q{row}'].value
    x = point_data_sh[f'T{row}'].value
    y = point_data_sh[f'U{row}'].value
    isRoad = True if point_data_sh[f'O{row}'].value is not None else False

    point_list.append(Point(point_index, x, y, isRoad, [], []))

    if point_data_sh[f'P{row}'].value is not None:
        red_point_index_list.append(point_index - 1)

print('1. point_list 불러오기 완료')

# 2. D_list 불러오기 및 설정
D_data_wb = load_workbook('./data/D_data.xlsx')
red_point_index = 0
for i in range(2, 3000 + 2):
    D_list = [[] for _ in range(6)]

    for time_type in range(6):
        D_data_sh = D_data_wb[D_data_wb.sheetnames[time_type]]
        D_list[time_type] = [D_data_sh[f'{chr(use_case)}{i}'].value for use_case in range(ord('D'), ord('H') + 1)]

    point_list[i - 2].D_list = D_list[:]

print('2.D_list 설정 완료')

# 3. distance_list 생성 및 설정
for point in point_list:
    distance_list = []

    for red_point_index in red_point_index_list:
        red_point = point_list[red_point_index]
        distance_list.append(round(haversine((point.y, point.x), (red_point.y, red_point.x), unit='m'), 2))

    point.distance_list = distance_list[:]

print('3. distance_list 설정 완료')

# 4. 크롤링 데이터를 불러와 possible_point_list, far_point_list 생성
crawling_data_wb = load_workbook('./data/raw_crawling_data.xlsx')
crawling_data_sh = crawling_data_wb.active

crawling_data_dict = {}
for row in crawling_data_sh.rows:
    red_point_index = row[0].value.split(' -> ')[0][1:-1]
    point_index = row[0].value.split(' -> ')[1][1:-1]
    crawling_distance = 0
    expected_time = -1
    is_too_close = False

    if 'routes' in row[1].value[0:10]: # walk
        crawling_distance = json.loads(row[1].value)['routes'][0]['summary']['distance']
        expected_time = json.loads(row[1].value)['routes'][0]['summary']['duration']

    elif '\"code\":0' in row[1].value[0:10]: # car
        crawling_distance = json.loads(row[1].value)['route']['3,0,0,0,0,0'][0]['summary']['distance']
        expected_time = round(json.loads(row[1].value)['route']['3,0,0,0,0,0'][0]['summary']['staticDuration'] / 1000, 2)

    elif '\"code\":3' : # walk - too close
        crawling_distance = 7.5
        expected_time = 15
        is_too_close = True

    elif '\"code\":1' : # car - too close
        crawling_distance = 7.5
        expected_time = 3
        is_too_close = True

    crawling_data_dict[f'{red_point_index}~{point_index}'] = (crawling_distance, expected_time, is_too_close)

print('4. possible_point_list, far_point_list 생성 완료')

# 5. red_point_list 생성
for i in range(170):
    red_point = point_list[red_point_index_list[i]]
    possible_point_list = []
    far_point_list = []

    for point in point_list:
        if point.point_index == red_point.point_index:
            continue

        if point.distance_list[i] <= 45:
            possible_point_list.append({
                'point_index': point.point_index,
                'D_list': point.D_list,
                'isRoad': point.isRoad,
                'distance': point.distance_list[i],
                'crawling_distance': crawling_data_dict[f'{red_point.point_index}~{point.point_index}'][0],
                'expected_time': crawling_data_dict[f'{red_point.point_index}~{point.point_index}'][1],
                'too_close': crawling_data_dict[f'{red_point.point_index}~{point.point_index}'][2]
            })
        elif 45 < point.distance_list[i] and point.distance_list[i] <= 130:
            far_point_list.append({
                'point_index': point.point_index,
                'D_list': point.D_list,
                'isRoad': point.isRoad,
                'distance': 100,
                'crawling_distance': -1,
                'expected_time': -1,
                'too_close': False
            })

    red_point_list.append(RedPoint(red_point.point_index, red_point.x, red_point.y, red_point.isRoad, red_point.D_list, red_point.distance_list, possible_point_list, far_point_list))

print('5. red_point_list 생성 완료')

# 6. point_list, red_point_list를 pickle 파일로 저장
with open('./data/new_point_list.p', 'wb') as f:
    for point in point_list:
        pickle.dump(point, f)

with open('./data/new_red_point_list.p', 'wb') as f:
    for red_point in red_point_list:
        pickle.dump(red_point, f)

print('6. pickle 파일로 저장 완료')
