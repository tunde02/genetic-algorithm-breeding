import requests, time, random
from haversine import haversine
from openpyxl import load_workbook, Workbook
from fake_useragent import UserAgent


def calc_distance(startX, startY, endX, endY):
    """
    경위도를 사용해서 (startX, startY) ~ (endX, endY) 거리 계산 후 반환
    haversine에는 (위도, 경도) 형태로 넘겨줘야됨 ex) (37.47, 127.04)
    """

    distance = haversine((startY, startX), (endY, endX), unit='m')

    return round(distance, 2)


def crawling(excel_file_path, save_file_name, start_column, end_column, max_Cz):
    point_list = []
    red_point_list = []

    workbook = load_workbook(excel_file_path, data_only=True)
    sh = workbook.active

    print('> Excel Loaded')

    # Load point_list & red_point_list
    for column in range(start_column, end_column + 1):
        point_index = sh[f'Q{column}'].value
        x = sh[f'T{column}'].value
        y = sh[f'U{column}'].value
        isRoad = True if sh[f'O{column}'].value is not None else False
        isRedPoint = True if sh[f'P{column}'].value is not None else False

        point_info = {
            'point_index': point_index,
            'x': x,
            'y': y,
            'isRoad': isRoad
        }

        point_list.append(point_info)

        if isRedPoint:
            red_point_list.append(point_info)

    workbook.close()
    print('> point_list & red_point_list Loaded')

    # Crawling
    try:
        save_wb = load_workbook(f'./data/{save_file_name}.xlsx')
    except:
        save_wb = Workbook()
    save_sh = save_wb.active

    crawling_count = 0
    for red_point in red_point_list:
        for point in point_list:
            # 최대 서비스 범위 밖의 점은 크롤링 대상에서 제외
            distance = calc_distance(red_point['x'], red_point['y'], point['x'], point['y'])
            if (distance > max_Cz) or (point is red_point):
                continue

            url = ''
            params = {}
            if point['isRoad']:
                url = 'https://map.naver.com/v5/api/dir/findcar'
                params = {
                    'start': f"{point['x']},{point['y']}",
                    'goal': f"{red_point['x']},{red_point['y']}",
                    'crs': 'EPSG:4326',
                    'mode': 'STATIC',
                    'lang': 'ko'
                }
            else:
                url = 'https://map.naver.com/v5/api/dir/findwalk'
                params = {
                    'lo': 'ko',
                    'r': 'step',
                    'st': '1',
                    'o': 'all',
                    'l': f"{point['x']},{point['y']};{red_point['x']},{red_point['y']}",
                    'lang': 'ko'
                }

            # 응답 받을때까지 시도
            with requests.Session() as session:
                while True:
                    crawling_count += 1

                    ua = UserAgent()
                    headers = { 'User-Agent': ua.random }
                    response = session.get(url, headers=headers, params=params)

                    if response.status_code == 200:
                        break

                    sleep_time = random.uniform(30, 60)
                    print(f"> Failed Crawling... ([{red_point['point_index']}] -> [{point['point_index']}]) (sleep {round(sleep_time, 1)}s, total {crawling_count})")
                    time.sleep(sleep_time)

            save_sh.append([f"[{red_point['point_index']}] -> [{point['point_index']}]", response.text])

            save_wb.save(f'./data/{save_file_name}.xlsx')

            sleep_time = random.uniform(1, 3)
            print(f"> Success Crawling! ([{point['point_index']}] ~ [{red_point['point_index']}]) (sleep {round(sleep_time, 1)}s, total {crawling_count})")
            time.sleep(sleep_time)

    save_wb.close()

    print('> Crawling Completed')


def get_expect_crawling_count(excel_file_path, start_column, end_column, max_Cz):
    point_list = []
    red_point_list = []
    total_crawling_count = 0

    workbook = load_workbook(excel_file_path, data_only=True)
    sh = workbook.active

    for column in range(start_column, end_column + 1):
        x = sh[f'T{column}'].value
        y = sh[f'U{column}'].value
        isRedPoint = True if sh[f'P{column}'].value is not None else False

        point_info = {'x': x, 'y': y}

        point_list.append(point_info)

        if isRedPoint:
            red_point_list.append(point_info)

    workbook.close()

    for red_point in red_point_list:
        for point in point_list:
            distance = calc_distance(red_point['x'], red_point['y'], point['x'], point['y'])
            if (distance > max_Cz) or (point is red_point):
                continue

            total_crawling_count += 1

    return total_crawling_count


if __name__ == "__main__":
    excel_file_path = './data/poi_coord.xlsx'
    save_file_name = 'test_raw_crawling_data'
    start_column = 2
    end_column = 31
    max_Cz = 45

    print(f"예상 크롤링 횟수 : {get_expect_crawling_count(excel_file_path, start_column, end_column, max_Cz)}")
    crawling(excel_file_path, save_file_name, start_column, end_column, max_Cz)
